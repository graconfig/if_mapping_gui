# tests/test_reader.py
import io
from pathlib import Path
import openpyxl
import pytest

from excel.reader import build_sheet_previews, PREVIEW_ROWS


def _make_xlsx(sheets: dict[str, list[list]]) -> Path:
    """Write an in-memory xlsx to a temp file and return its path."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for r, row in enumerate(rows, 1):
            for c, val in enumerate(row, 1):
                ws.cell(r, c, val)
    import tempfile, os
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    return Path(path)


def test_preview_contains_sheet_name():
    path = _make_xlsx({"Sheet1": [["hello", "world"]]})
    previews = build_sheet_previews(path)
    assert len(previews) == 1
    assert previews[0]["sheetName"] == "Sheet1"
    path.unlink()


def test_preview_text_format():
    """Each cell must appear as [ColLetter]value."""
    path = _make_xlsx({"Data": [["alpha", "beta"]]})
    previews = build_sheet_previews(path)
    text = previews[0]["previewText"]
    assert "row1:" in text
    assert "[A]alpha" in text
    assert "[B]beta" in text
    path.unlink()


def test_preview_row_limit():
    """Only first PREVIEW_ROWS rows should be included."""
    rows = [[f"r{i}"] for i in range(PREVIEW_ROWS + 5)]
    path = _make_xlsx({"Big": rows})
    previews = build_sheet_previews(path)
    line_count = previews[0]["previewText"].count("\n") + 1
    assert line_count <= PREVIEW_ROWS
    path.unlink()


def test_empty_sheet_skipped():
    """Sheets with no non-empty cells should be excluded."""
    path = _make_xlsx({"Empty": [[None, None]], "Real": [["x"]]})
    previews = build_sheet_previews(path)
    names = [p["sheetName"] for p in previews]
    assert "Empty" not in names
    assert "Real" in names
    path.unlink()


def test_multiple_sheets():
    path = _make_xlsx({
        "対象IF":       [["module", "MyMod"]],
        "IFマッピング定義": [["fieldName", "tableId"]],
    })
    previews = build_sheet_previews(path)
    assert len(previews) == 2
    path.unlink()
