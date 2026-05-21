# tests/test_excel_writer.py
import openpyxl
import pytest
from pathlib import Path
from excel.writer import write_results
import config


def _default_output_cols(direction: str = "normal") -> dict:
    return config.load()["excel"]["directions"][direction]["output_cols"]


def _make_workbook(sheet_head: str = "対属IF", sheet_data: str = "IFマッピング定義",
                   data_rows: list[dict] | None = None) -> openpyxl.Workbook:
    cfg = config.load()["excel"]
    sheet_head = cfg["sheet_head"]
    sheet_data = cfg["sheet_data"]
    start = cfg["start_row"]
    wb = openpyxl.Workbook()
    wb.active.title = sheet_head
    ws = wb.create_sheet(sheet_data)
    if data_rows:
        for i, row in enumerate(data_rows):
            ws[f"C{start + i}"] = row.get("fieldName", "")
    return wb


def test_output_filename_format(tmp_path):
    output_cols = _default_output_cols()
    wb = _make_workbook(data_rows=[{"fieldName": "MATNR"}])
    input_path = tmp_path / "IF_MM_001.xlsx"
    start = config.load()["excel"]["start_row"]
    results = [{"rowIndex": start, "tableId": "T", "fieldId": "F",
                "dataType": "Char", "fieldText": "Material", "matchScore": 1.0,
                "matchSource": "custom", "notes": ""}]
    out = write_results(input_path, wb, results, output_cols)
    assert out.name.startswith("processed_")
    assert out.name.endswith("IF_MM_001.xlsx")
    assert out.parent == tmp_path


def test_writes_to_correct_columns(tmp_path):
    output_cols = _default_output_cols()
    wb = _make_workbook(data_rows=[{"fieldName": "MATNR"}])
    input_path = tmp_path / "test.xlsx"
    start = config.load()["excel"]["start_row"]
    results = [{
        "rowIndex": start,
        "tableId": "CDS_View", "fieldId": "Material",
        "dataType": "Char10", "fieldText": "品目コード",
        "matchScore": 0.95, "matchSource": "ai", "notes": "High confidence",
    }]
    out = write_results(input_path, wb, results, output_cols)

    sheet_data = config.load()["excel"]["sheet_data"]
    saved_ws = openpyxl.load_workbook(out)[sheet_data]
    assert saved_ws[f"{output_cols['table_id']}{start}"].value == "CDS_View"
    assert saved_ws[f"{output_cols['field_id']}{start}"].value == "Material"
    assert saved_ws[f"{output_cols['data_type']}{start}"].value == "Char10"
    assert saved_ws[f"{output_cols['field_name']}{start}"].value == "品目コード"
    assert saved_ws[f"{output_cols['match_source']}{start}"].value == "ai"
    assert saved_ws[f"{output_cols['notes']}{start}"].value == "High confidence"


def test_result_matched_by_rowindex(tmp_path):
    output_cols = _default_output_cols()
    wb = _make_workbook(data_rows=[{"fieldName": "MATNR"}, {"fieldName": "WERKS"}])
    input_path = tmp_path / "test.xlsx"
    start = config.load()["excel"]["start_row"]
    row1, row2 = start, start + 1
    results = [
        {"rowIndex": row2, "tableId": "T2", "fieldId": "F2",
         "dataType": "", "fieldText": "", "matchScore": 0.8, "matchSource": "ai", "notes": ""},
        {"rowIndex": row1, "tableId": "T1", "fieldId": "F1",
         "dataType": "", "fieldText": "", "matchScore": 1.0, "matchSource": "custom", "notes": ""},
    ]
    out = write_results(input_path, wb, results, output_cols)
    sheet_data = config.load()["excel"]["sheet_data"]
    ws = openpyxl.load_workbook(out)[sheet_data]
    assert ws[f"{output_cols['table_id']}{row1}"].value == "T1"
    assert ws[f"{output_cols['table_id']}{row2}"].value == "T2"


def test_rows_with_no_result_unchanged(tmp_path):
    output_cols = _default_output_cols()
    wb = _make_workbook(data_rows=[{"fieldName": "MATNR"}, {"fieldName": "WERKS"}])
    input_path = tmp_path / "test.xlsx"
    start = config.load()["excel"]["start_row"]
    row1, row2 = start, start + 1
    results = [{"rowIndex": row1, "tableId": "T1", "fieldId": "F1",
                "dataType": "", "fieldText": "", "matchScore": 1.0, "matchSource": "custom", "notes": ""}]
    out = write_results(input_path, wb, results, output_cols)
    sheet_data = config.load()["excel"]["sheet_data"]
    ws = openpyxl.load_workbook(out)[sheet_data]
    assert ws[f"{output_cols['table_id']}{row2}"].value is None


def test_sap_direction_writes_to_sap_output_cols(tmp_path):
    output_cols = _default_output_cols("sap")
    wb = _make_workbook(data_rows=[{"fieldName": "MATNR"}])
    input_path = tmp_path / "test_sap.xlsx"
    start = config.load()["excel"]["start_row"]
    results = [{"rowIndex": start, "tableId": "EKKO", "fieldId": "EBELN",
                "dataType": "CHAR10", "fieldText": "PO Number",
                "matchScore": 0.9, "matchSource": "ai", "notes": ""}]
    out = write_results(input_path, wb, results, output_cols)
    sheet_data = config.load()["excel"]["sheet_data"]
    ws = openpyxl.load_workbook(out)[sheet_data]
    # SAP direction: table_id → G, field_id → H
    assert ws[f"G{start}"].value == "EKKO"
    assert ws[f"H{start}"].value == "EBELN"
