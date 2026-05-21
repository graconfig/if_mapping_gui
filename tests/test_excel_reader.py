# tests/test_excel_reader.py
import pytest
import openpyxl
from pathlib import Path
from excel.reader import read_fields, read_kb_fields, InterfaceFieldInput, ExcelReadError
import config


def _default_excel_cfg() -> dict:
    return config.load()["excel"]


def _make_if_xlsx(
    tmp_path: Path,
    *,
    module: str = "MM",
    if_name: str = "TEST_IF",
    if_desc: str = "Test interface",
    data_rows: list[dict] | None = None,
    detection_value: str = "",
    excel_cfg: dict | None = None,
) -> Path:
    cfg = excel_cfg or _default_excel_cfg()
    sheet_head = cfg["sheet_head"]
    sheet_data = cfg["sheet_data"]
    header_row = cfg["header_row"]

    direction = "sap" if detection_value and cfg["detection"]["keyword"].upper() in detection_value.upper() else "normal"
    hcols = cfg["directions"][direction]["input_header_cols"]
    rcols = cfg["directions"][direction]["input_row_cols"]
    start_row = cfg["start_row"]
    det = cfg["detection"]

    wb = openpyxl.Workbook()
    wb.active.title = sheet_head
    ws_head = wb[sheet_head]
    ws_head[f"{hcols['module']}{header_row}"] = module
    ws_head[f"{hcols['if_name']}{header_row}"] = if_name
    ws_head[f"{hcols['if_desc']}{header_row}"] = if_desc
    if detection_value:
        ws_head[f"{det['col']}{det['row']}"] = detection_value

    ws_data = wb.create_sheet(sheet_data)
    if data_rows is None:
        data_rows = [{"field_name": "MATNR", "field_text": "品目コード", "sample_value": "100", "remark": ""}]
    for i, r in enumerate(data_rows):
        row_num = start_row + i
        for col_key, col_letter in rcols.items():
            if col_key in r:
                ws_data[f"{col_letter}{row_num}"] = r[col_key]

    path = tmp_path / "test_if.xlsx"
    wb.save(path)
    return path


# ── read_fields: normal direction ────────────────────────────────────────────

def test_reads_interface_fields(tmp_path):
    excel_cfg = _default_excel_cfg()
    path = _make_if_xlsx(tmp_path, excel_cfg=excel_cfg, data_rows=[
        {"field_name": "MATNR", "field_text": "品目コード", "sample_value": "100", "remark": "備考"},
        {"field_name": "WERKS", "field_text": "プラント",   "sample_value": "1000", "remark": ""},
    ])
    fields, wb, direction = read_fields(path, excel_cfg)
    assert direction == "normal"
    assert len(fields) == 2
    assert fields[0].fieldName == "MATNR"
    assert fields[0].fieldText == "品目コード"
    assert fields[0].sampleValue == "100"
    assert fields[0].remark == "備考"
    assert fields[1].fieldName == "WERKS"
    wb.close()


def test_reads_header_from_head_sheet(tmp_path):
    excel_cfg = _default_excel_cfg()
    path = _make_if_xlsx(tmp_path, excel_cfg=excel_cfg, module="SD", if_name="IF_SD_001", if_desc="Sales order")
    fields, wb, direction = read_fields(path, excel_cfg)
    assert fields[0].module == "SD"
    assert fields[0].ifName == "IF_SD_001"
    assert fields[0].ifDesc == "Sales order"
    wb.close()


def test_rowindex_matches_sheet_row(tmp_path):
    excel_cfg = _default_excel_cfg()
    start = excel_cfg["start_row"]
    path = _make_if_xlsx(tmp_path, excel_cfg=excel_cfg, data_rows=[
        {"field_name": "MATNR"}, {"field_name": "WERKS"},
    ])
    fields, wb, _ = read_fields(path, excel_cfg)
    assert fields[0].rowIndex == start
    assert fields[1].rowIndex == start + 1
    wb.close()


def test_skips_blank_rows(tmp_path):
    excel_cfg = _default_excel_cfg()
    path = _make_if_xlsx(tmp_path, excel_cfg=excel_cfg, data_rows=[
        {"field_name": "MATNR"}, {"field_name": None}, {"field_name": ""}, {"field_name": "WERKS"},
    ])
    fields, wb, _ = read_fields(path, excel_cfg)
    assert len(fields) == 2
    assert fields[1].fieldName == "WERKS"
    wb.close()


def test_raises_on_missing_head_sheet(tmp_path):
    excel_cfg = _default_excel_cfg()
    wb = openpyxl.Workbook()
    wb.active.title = excel_cfg["sheet_data"]
    path = tmp_path / "no_head.xlsx"
    wb.save(path)
    with pytest.raises(ExcelReadError, match=excel_cfg["sheet_head"]):
        read_fields(path, excel_cfg)


def test_raises_on_missing_data_sheet(tmp_path):
    excel_cfg = _default_excel_cfg()
    wb = openpyxl.Workbook()
    wb.active.title = excel_cfg["sheet_head"]
    path = tmp_path / "no_data.xlsx"
    wb.save(path)
    with pytest.raises(ExcelReadError, match=excel_cfg["sheet_data"]):
        read_fields(path, excel_cfg)


def test_to_dict_includes_all_new_fields(tmp_path):
    excel_cfg = _default_excel_cfg()
    path = _make_if_xlsx(tmp_path, excel_cfg=excel_cfg, data_rows=[{
        "field_name": "MATNR", "field_text": "品目", "sample_value": "X", "remark": "R",
        "table_id": "EKKO", "field_id": "EBELN", "key_flag": "○", "obligatory": "必須",
        "data_type": "CHAR", "length_total": "10", "length_dec": "0",
        "is_append": "Y", "verify": "○",
    }])
    fields, wb, _ = read_fields(path, excel_cfg)
    d = fields[0].to_dict()
    expected_keys = {
        "rowIndex", "module", "ifName", "ifDesc",
        "fieldName", "fieldText", "sampleValue", "remark",
        "tableId", "fieldId", "keyFlag", "obligatory",
        "dataType", "lengthTotal", "lengthDec", "isAppend", "verify",
    }
    assert set(d.keys()) == expected_keys
    assert d["tableId"] == "EKKO"
    assert d["fieldId"] == "EBELN"
    wb.close()


# ── direction detection ───────────────────────────────────────────────────────

def test_detects_sap_direction(tmp_path):
    excel_cfg = _default_excel_cfg()
    path = _make_if_xlsx(tmp_path, excel_cfg=excel_cfg, detection_value="SAP S/4HANA",
                          data_rows=[{"field_name": "MATNR"}])
    fields, wb, direction = read_fields(path, excel_cfg)
    assert direction == "sap"
    wb.close()


def test_detects_normal_direction_when_no_keyword(tmp_path):
    excel_cfg = _default_excel_cfg()
    path = _make_if_xlsx(tmp_path, excel_cfg=excel_cfg, detection_value="External System",
                          data_rows=[{"field_name": "MATNR"}])
    fields, wb, direction = read_fields(path, excel_cfg)
    assert direction == "normal"
    wb.close()


def test_detects_normal_direction_when_detection_cell_empty(tmp_path):
    excel_cfg = _default_excel_cfg()
    path = _make_if_xlsx(tmp_path, excel_cfg=excel_cfg, data_rows=[{"field_name": "MATNR"}])
    fields, wb, direction = read_fields(path, excel_cfg)
    assert direction == "normal"
    wb.close()


def test_sap_direction_reads_from_sap_columns(tmp_path):
    excel_cfg = _default_excel_cfg()
    path = _make_if_xlsx(tmp_path, excel_cfg=excel_cfg, detection_value="SAP",
                          data_rows=[{"field_name": "MATNR_SAP", "field_text": "SAP品目"}])
    fields, wb, direction = read_fields(path, excel_cfg)
    assert direction == "sap"
    assert fields[0].fieldName == "MATNR_SAP"
    assert fields[0].fieldText == "SAP品目"
    wb.close()


# ── read_kb_fields (unchanged) ────────────────────────────────────────────────

def _make_kb_xlsx(tmp_path: Path, headers: list, rows: list[list]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    path = tmp_path / "kb.xlsx"
    wb.save(path)
    return path


def test_kb_reads_standard_columns(tmp_path):
    path = _make_kb_xlsx(tmp_path,
        ["sourceField", "sourceDesc", "sourceTable", "targetField", "targetTable"],
        [["MATNR", "品目コード", "EKPO", "Material", "C_PurchaseOrderItemTP"]],
    )
    records = read_kb_fields(path)
    assert len(records) == 1
    assert records[0]["sourceField"] == "MATNR"
    assert records[0]["targetField"] == "Material"


def test_kb_skips_blank_rows(tmp_path):
    path = _make_kb_xlsx(tmp_path,
        ["sourceField", "sourceDesc"],
        [["MATNR", "品目"], [None, None], ["WERKS", "プラント"]],
    )
    records = read_kb_fields(path)
    assert len(records) == 2


def test_kb_raises_on_missing_sourcefield(tmp_path):
    path = _make_kb_xlsx(tmp_path, ["sourceDesc"], [["Material"]])
    with pytest.raises(ExcelReadError, match="sourceField"):
        read_kb_fields(path)
