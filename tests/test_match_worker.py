# tests/test_match_worker.py
import queue
import threading
from pathlib import Path
from unittest.mock import patch
import openpyxl
import pytest

from gui.frames.match_frame import match_worker
import config


def _default_excel_cfg() -> dict:
    return config.load()["excel"]


def _make_if_xlsx(tmp_path: Path, name: str = "input.xlsx",
                  detection_value: str = "") -> Path:
    cfg = _default_excel_cfg()
    sheet_head = cfg["sheet_head"]
    sheet_data = cfg["sheet_data"]
    header_row = cfg["header_row"]
    start_row = cfg["start_row"]
    hcols = cfg["directions"]["normal"]["input_header_cols"]
    rcols = cfg["directions"]["normal"]["input_row_cols"]
    det = cfg["detection"]

    wb = openpyxl.Workbook()
    wb.active.title = sheet_head
    ws_head = wb[sheet_head]
    ws_head[f"{hcols['if_name']}{header_row}"] = "IF_TEST"
    ws_head[f"{hcols['module']}{header_row}"] = "MM"
    ws_head[f"{hcols['if_desc']}{header_row}"] = "Test"
    if detection_value:
        ws_head[f"{det['col']}{det['row']}"] = detection_value
    ws_data = wb.create_sheet(sheet_data)
    ws_data[f"{rcols['field_name']}{start_row}"] = "MATNR"
    ws_data[f"{rcols['field_text']}{start_row}"] = "品目コード"
    path = tmp_path / name
    wb.save(path)
    return path


def _drain(q: queue.Queue) -> list[dict]:
    msgs = []
    while not q.empty():
        msgs.append(q.get_nowait())
    return msgs


def test_worker_emits_done_with_results(tmp_path):
    xlsx = _make_if_xlsx(tmp_path)
    excel_cfg = _default_excel_cfg()
    q = queue.Queue()
    stop = threading.Event()
    start = excel_cfg["start_row"]
    mock_results = [{"rowIndex": start, "tableId": "C_PO", "fieldId": "Material",
                     "matchSource": "custom", "matchScore": 1.0, "notes": ""}]
    with patch("gui.frames.match_frame.CapClient") as MockClient:
        MockClient.return_value.ping.return_value = True
        MockClient.return_value.match.return_value = mock_results
        match_worker([str(xlsx)], "claude", "ja", "http://localhost:4004", excel_cfg, q, stop)
    msgs = _drain(q)
    assert "done" in [m["type"] for m in msgs]
    done_msg = next(m for m in msgs if m["type"] == "done")
    assert done_msg["results"] == mock_results


def test_worker_stores_direction_in_done_message(tmp_path):
    xlsx_normal = _make_if_xlsx(tmp_path, name="normal.xlsx")
    xlsx_sap    = _make_if_xlsx(tmp_path, name="sap.xlsx", detection_value="SAP")
    excel_cfg = _default_excel_cfg()
    q = queue.Queue()
    stop = threading.Event()
    with patch("gui.frames.match_frame.CapClient") as MockClient:
        MockClient.return_value.ping.return_value = True
        MockClient.return_value.match.return_value = []
        match_worker([str(xlsx_normal), str(xlsx_sap)], "claude", "ja",
                     "http://localhost:4004", excel_cfg, q, stop)
    msgs = _drain(q)
    done = next(m for m in msgs if m["type"] == "done")
    assert done["directions"][str(xlsx_normal)] == "normal"
    assert done["directions"][str(xlsx_sap)] == "sap"


def test_worker_emits_error_when_cap_unreachable(tmp_path):
    xlsx = _make_if_xlsx(tmp_path)
    excel_cfg = _default_excel_cfg()
    q = queue.Queue()
    stop = threading.Event()
    with patch("gui.frames.match_frame.CapClient") as MockClient:
        MockClient.return_value.ping.return_value = False
        match_worker([str(xlsx)], "claude", "ja", "http://localhost:4004", excel_cfg, q, stop)
    msgs = _drain(q)
    assert any(m["type"] == "error" for m in msgs)


def test_worker_skips_bad_excel_and_continues(tmp_path):
    bad = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "WrongSheet"
    wb.save(bad)

    good = _make_if_xlsx(tmp_path, name="good.xlsx")
    excel_cfg = _default_excel_cfg()
    q = queue.Queue()
    stop = threading.Event()
    with patch("gui.frames.match_frame.CapClient") as MockClient:
        MockClient.return_value.ping.return_value = True
        MockClient.return_value.match.return_value = []
        match_worker([str(bad), str(good)], "claude", "ja", "http://localhost:4004", excel_cfg, q, stop)
    msgs = _drain(q)
    log_texts = [m["text"] for m in msgs if m["type"] == "log"]
    assert any("ERROR" in t for t in log_texts)
    assert any(m["type"] == "done" for m in msgs)


def test_worker_respects_stop_event(tmp_path):
    files = []
    for i in range(3):
        sub = tmp_path / f"f{i}"
        sub.mkdir()
        files.append(str(_make_if_xlsx(sub, name="input.xlsx")))
    excel_cfg = _default_excel_cfg()
    q = queue.Queue()
    stop = threading.Event()
    stop.set()
    with patch("gui.frames.match_frame.CapClient") as MockClient:
        MockClient.return_value.ping.return_value = True
        match_worker(files, "claude", "ja", "http://localhost:4004", excel_cfg, q, stop)
    msgs = _drain(q)
    log_texts = [m["text"] for m in msgs if m["type"] == "log"]
    assert any("停止" in t for t in log_texts)
