# excel/writer.py
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

# Maps CAP result camelCase keys → output_cols snake_case keys
_RESULT_KEY_MAP: dict[str, str] = {
    "fieldText":   "field_name",
    "tableId":     "table_id",
    "fieldId":     "field_id",
    "dataType":    "data_type",
    "lengthTotal": "length_total",
    "lengthDec":   "length_dec",
    "keyFlag":     "key_flag",
    "obligatory":  "obligatory",
    "notes":       "notes",
    "sampleValue": "sample_value",
    "matchScore":  "match_score",
    "matchSource": "match_source",
}

_MATCH_SOURCE_LABELS: dict[str, str] = {
    "exact":  "対応表マッピング",
    "vector": "対応表マッピング",
    "ai":     "AIマッピング",
}


def _transform(cap_key: str, value: object) -> object:
    if cap_key == "matchScore" and value is not None:
        return round(float(value) * 100)  # type: ignore[arg-type]
    if cap_key == "matchSource":
        return _MATCH_SOURCE_LABELS.get(str(value), str(value))
    return value


def write_results(
    input_path: Path,
    workbook: openpyxl.Workbook,
    results: list[dict],
    output_cols: dict[str, str],
    sheet_data: str = "IFマッピング定義",
    input_row_cols: dict[str, str] | None = None,
    output_dir: Path | None = None,
) -> Path:
    """Write match results into workbook's data sheet, save as processed_*.xlsx.

    output_cols maps snake_case field names to Excel column letters,
    e.g. {"table_id": "W", "field_id": "X", ...}.
    input_row_cols maps input column roles to column letters,
    e.g. {"verify": "C", "is_append": "D"}.
    """
    ws = workbook[sheet_data]
    result_map: dict[int, dict] = {r["rowIndex"]: r for r in results if "rowIndex" in r}

    verify_col       = (input_row_cols or {}).get("verify")
    is_append_col_in = (input_row_cols or {}).get("is_append")
    is_append_col_out = output_cols.get("is_append")

    table_id_col = output_cols.get("table_id")
    field_id_col = output_cols.get("field_id")

    for row_idx, result in result_map.items():
        if verify_col and ws[f"{verify_col}{row_idx}"].value == "○":
            continue

        for cap_key, col_key in _RESULT_KEY_MAP.items():
            if cap_key not in result or col_key not in output_cols:
                continue
            col_letter = output_cols[col_key]
            ws[f"{col_letter}{row_idx}"] = _transform(cap_key, result[cap_key])

        if is_append_col_in and is_append_col_out:
            ws[f"{is_append_col_out}{row_idx}"] = ws[f"{is_append_col_in}{row_idx}"].value

        color_code = result.get("color", "")
        if color_code and result.get("matchSource") in ("exact", "vector"):
            hex_color = color_code.lstrip("#").upper()
            if len(hex_color) == 6:
                fill = PatternFill(fill_type="solid", fgColor=hex_color)
                for col in filter(None, [table_id_col, field_id_col]):
                    ws[f"{col}{row_idx}"].fill = fill

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = output_dir if output_dir is not None else input_path.parent
    out_path = dest / f"processed_{ts}_{input_path.name}"
    workbook.save(out_path)
    return out_path
