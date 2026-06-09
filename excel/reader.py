# excel/reader.py
import warnings
import openpyxl
from pathlib import Path
from dataclasses import dataclass
from openpyxl.utils import get_column_letter

warnings.filterwarnings(
    "ignore",
    message="DrawingML support is incomplete.*",
    category=UserWarning,
)

PREVIEW_ROWS = 15


def build_sheet_previews(path: Path) -> list[dict]:
    """Return [{sheetName, previewText}] for all non-empty sheets.

    previewText: first PREVIEW_ROWS rows, each cell as [ColLetter]value,
    cells separated by spaces, rows separated by newlines.
    Empty sheets (all preview rows are None) are skipped.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        raise ExcelReadError(f"{path.name}: cannot open for preview — {e}") from e

    previews: list[dict] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines: list[str] = []
            for row in ws.iter_rows(max_row=PREVIEW_ROWS, values_only=False):
                cells = []
                for cell in row:
                    if cell.value is not None:
                        col_letter = get_column_letter(cell.column)
                        cells.append(f"[{col_letter}]{cell.value}")
                if cells:
                    lines.append(" ".join(cells))
            if lines:
                previews.append({
                    "sheetName":   sheet_name,
                    "previewText": "\n".join(lines),
                })
    finally:
        wb.close()
    return previews


# KB_FIELDS defines the canonical field order for header-based parsing fallback
KB_FIELDS = ["ifName", "sourceDesc", "sourceTable", "sourceField",
             "targetDesc", "targetTable", "targetField", "notes"]

KB_ALIASES: dict[str, str] = {
    "ifname":      "ifName",
    "sourcedesc":  "sourceDesc",
    "sourcetable": "sourceTable",
    "sourcefield": "sourceField",
    "targetdesc":  "targetDesc",
    "targettable": "targetTable",
    "targetfield": "targetField",
    "notes":       "notes",
}

# Maps KB_FIELDS canonical names → column keys used in kb_upload config
_KB_FIELD_TO_COL_KEY: dict[str, str] = {
    "ifName":      "if_name",
    "sourceDesc":  "source_desc",
    "sourceTable": "source_table",
    "sourceField": "source_field",
    "targetDesc":  "target_desc",
    "targetTable": "target_table",
    "targetField": "target_field",
    "notes":       "notes",
}


@dataclass
class InterfaceFieldInput:
    rowIndex:    int
    module:      str = ""
    ifName:      str = ""
    ifDesc:      str = ""
    fieldName:   str = ""
    fieldText:   str = ""
    sampleValue: str = ""
    remark:      str = ""
    tableId:     str = ""
    fieldId:     str = ""
    keyFlag:     str = ""
    obligatory:  str = ""
    dataType:    str = ""
    lengthTotal: str = ""
    lengthDec:   str = ""
    isAppend:    str = ""
    verify:      str = ""

    def to_dict(self) -> dict:
        return {
            "rowIndex":    self.rowIndex,
            "module":      self.module,
            "ifName":      self.ifName,
            "ifDesc":      self.ifDesc,
            "fieldName":   self.fieldName,
            "fieldText":   self.fieldText,
            "sampleValue": self.sampleValue,
            "remark":      self.remark,
            "tableId":     self.tableId,
            "fieldId":     self.fieldId,
            "keyFlag":     self.keyFlag,
            "obligatory":  self.obligatory,
            "dataType":    self.dataType,
            "lengthTotal": self.lengthTotal,
            "lengthDec":   self.lengthDec,
            "isAppend":    self.isAppend,
            "verify":      self.verify,
        }


class ExcelReadError(Exception):
    pass


def _cell_str(ws, col: str, row: int) -> str:
    v = ws[f"{col}{row}"].value
    return str(v).strip() if v is not None else ""


def read_fields(
    path: Path, excel_cfg: dict
) -> tuple[list[InterfaceFieldInput], openpyxl.Workbook, str]:
    """Read an IF mapping Excel using excel_cfg column definitions.

    Returns (fields, workbook, direction) where direction is 'normal' or 'sap'.
    The workbook is kept open so write_results() can use it.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise ExcelReadError(f"{path.name}: cannot open — {e}") from e

    sheet_head = excel_cfg["sheet_head"]
    sheet_data = excel_cfg["sheet_data"]

    if sheet_head not in wb.sheetnames:
        raise ExcelReadError(
            f"{path.name}: sheet '{sheet_head}' not found (sheets: {wb.sheetnames})"
        )
    if sheet_data not in wb.sheetnames:
        raise ExcelReadError(
            f"{path.name}: sheet '{sheet_data}' not found (sheets: {wb.sheetnames})"
        )

    ws_head = wb[sheet_head]
    ws_data = wb[sheet_data]

    # Detect direction
    det = excel_cfg["detection"]
    det_value = _cell_str(ws_head, det["col"], det["row"])
    keyword = det["keyword"].upper()
    direction = "sap" if keyword in det_value.upper() else "normal"

    dir_cfg = excel_cfg["directions"][direction]
    hcols = dir_cfg["input_header_cols"]
    rcols = dir_cfg["input_row_cols"]
    header_row = excel_cfg["header_row"]
    start_row = excel_cfg["start_row"]

    module  = _cell_str(ws_head, hcols["module"],  header_row)
    if_name = _cell_str(ws_head, hcols["if_name"], header_row)
    if_desc = _cell_str(ws_head, hcols["if_desc"], header_row)

    fields: list[InterfaceFieldInput] = []
    max_row = ws_data.max_row or 1000

    for row in range(start_row, max_row + 1):
        fn_col = rcols["field_name"]
        field_name_raw = ws_data[f"{fn_col}{row}"].value
        if field_name_raw is None or str(field_name_raw).strip() in ("", "e"):
            continue

        def _get(key: str, _row: int = row) -> str:
            col = rcols.get(key)
            return _cell_str(ws_data, col, _row) if col else ""

        fields.append(InterfaceFieldInput(
            rowIndex=row,
            module=module,
            ifName=if_name,
            ifDesc=if_desc,
            fieldName=str(field_name_raw).strip(),
            fieldText=_get("field_text"),
            sampleValue=_get("sample_value"),
            remark=_get("remark"),
            tableId=_get("table_id"),
            fieldId=_get("field_id"),
            keyFlag=_get("key_flag"),
            obligatory=_get("obligatory"),
            dataType=_get("data_type"),
            lengthTotal=_get("length_total"),
            lengthDec=_get("length_dec"),
            isAppend=_get("is_append"),
            verify=_get("verify"),
        ))

    return fields, wb, direction


def _cell_hex_color(cell) -> str:
    """Extract hex color (#RRGGBB) from a cell's background fill, or empty string."""
    try:
        fill = cell.fill
        if fill and fill.fill_type not in (None, "none"):
            fg = fill.fgColor
            if fg.type == "rgb":
                argb = fg.rgb
                if argb and argb != "00000000":
                    return "#" + argb[-6:]
    except Exception:
        pass
    return ""


def read_kb_fields(path: Path, kb_cfg: dict | None = None) -> list[dict]:
    """Read a knowledge-base Excel → list of CustomFieldUploadInput dicts.

    Uses kb_cfg (from config["kb_upload"]) for sheet name, start row, and
    column positions.  Falls back to header-name detection when kb_cfg is None.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise ExcelReadError(f"{path.name}: cannot open — {e}") from e

    if kb_cfg:
        # ── config-driven (positional columns) ────────────────────────────
        sheet_name  = kb_cfg.get("sheet_name")
        start_row   = int(kb_cfg.get("start_row", 2))
        skip_value  = kb_cfg.get("skip_value", "")
        col_cfg     = kb_cfg.get("columns", {})
        color_key   = kb_cfg.get("color_column", "target_desc")
        color_idx   = int(col_cfg.get(color_key, 0)) - 1  # 0-based

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            preferred = [s for s in wb.sheetnames if "正本" in s]
            ws = wb[preferred[0]] if preferred else wb.active

        def _col(key: str):
            idx = col_cfg.get(key)
            return int(idx) - 1 if idx is not None else None   # 0-based

        records: list[dict] = []
        for row_cells in ws.iter_rows(min_row=start_row):
            row = [c.value for c in row_cells]

            def _v(key: str) -> str:
                i = _col(key)
                if i is None or i >= len(row):
                    return ""
                v = row[i]
                return str(v).strip() if v is not None else ""

            source_desc = _v("source_desc")
            if not source_desc or source_desc in ("", skip_value):
                continue

            color = ""
            if 0 <= color_idx < len(row_cells):
                color = _cell_hex_color(row_cells[color_idx])

            records.append({
                "ifName":      _v("if_name"),
                "sourceDesc":  source_desc,
                "sourceTable": _v("source_table"),
                "sourceField": _v("source_field"),
                "targetDesc":  _v("target_desc"),
                "targetTable": _v("target_table"),
                "targetField": _v("target_field"),
                "notes":       _v("notes"),
                "color":       color,
            })

        wb.close()
        return records

    # ── header-detection fallback (original behaviour) ────────────────────
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise ExcelReadError(f"{path.name}: empty file")

    headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        canonical = KB_ALIASES.get(h.lower(), h.lower())
        col_map[canonical] = i

    if "sourceField" not in col_map:
        raise ExcelReadError(
            f"{path.name}: missing required column 'sourceField' (got: {headers})"
        )

    records = []
    for row in all_rows[1:]:
        idx = col_map["sourceField"]
        sf = row[idx] if idx < len(row) else None
        if not sf:
            continue
        record: dict = {}
        for field in KB_FIELDS:
            i = col_map.get(field)
            record[field] = str(row[i]).strip() if i is not None and i < len(row) and row[i] else ""
        records.append(record)

    return records
